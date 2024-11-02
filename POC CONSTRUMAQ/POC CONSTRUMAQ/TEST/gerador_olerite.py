from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import mm
from io import BytesIO
from gerar_sub_total_um import Sub_total_um
from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
import os

class Gerar_olerite:
    def __init__(self, funcao):
        self.funcao = funcao
        self.margin = 20  # Margens em mm
        self.page_width, self.page_height = A4  

    def gerar_sub_um(self):
        total_pagamento = Sub_total_um.calcular_pagamento_um(self.funcao)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=0,  # Margem esquerda
                                rightMargin=0,  # Margem direita
                                topMargin=30,  # Margem superior
                                bottomMargin=30)  # Margem inferior

        # Cabeçalho
        elements = []
        self._draw_header(elements)

        # Conteúdo
        content = self._prepare_content(total_pagamento)
        table = self._create_table(content)
        elements.append(table)

        # Rodapé
        self._draw_footer(elements)

        doc.build(elements)
        buffer.seek(0)
        
        
        # Dados a serem exportados para Excel
        nome = self.funcao.funcionario['nome_funcionario']
        cpf = self.funcao.funcionario['numero_cpf']
        chave_pix = self.funcao.funcionario['chave_pix']
        valor_total = round(total_pagamento['sub_total_tres'], 2) 
        date_pagamento = self.funcao.data_pagamento.strftime("%d/%m/%Y")
        
        # Diretório e nome do arquivo Excel
        diretorio = 'static'
        arquivo_excel = os.path.join(diretorio, 'dados_de_pagamento.xlsx')
        
        # Cria o diretório "static" se ele não existir
        if not os.path.exists(diretorio):
            os.makedirs(diretorio)
        
        # Verifica se o arquivo já existe
        if os.path.exists(arquivo_excel):
            # Carrega o arquivo existente
            workbook = load_workbook(arquivo_excel)
            sheet = workbook.active
            
            # Verifica se o nome já existe e substitui os dados, se necessário
            for row in sheet.iter_rows(min_row=2):  # Ignora o cabeçalho
                if row[0].value == nome:  # Verifica se o nome já está presente
                    # Substitui a linha existente
                    row[1].value = cpf
                    row[2].value = chave_pix
                    row[3].value = valor_total
                    row[4].value = date_pagamento
                    break
            else:
                # Se o nome não foi encontrado, adiciona uma nova linha
                sheet.append([nome, cpf, chave_pix, valor_total, date_pagamento])

        else:
            # Se o arquivo não existir, cria um novo
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Dados do Funcionário"
            
            # Adiciona cabeçalhos
            headers = ["Nome", "CPF", "Chave PIX", "Valor Total", "Data de Pagamento"]
            sheet.append(headers)
            
            # Adiciona novos dados na planilha
            data = [nome, cpf, chave_pix, valor_total, date_pagamento]
            sheet.append(data)

        # Salva o arquivo Excel
        workbook.save(arquivo_excel)
        return buffer

    def _draw_header(self, elements):
        total_pagamento = self.funcao.calcular_pagamento_um()

        valor_sub_total_tres = total_pagamento.get('sub_total_tres')
      
        
        
        header = [
            ["VR3 LTDA", "", ""],
            ["CNPJ:12.507.345/0001-15", "", ""],
            ["", "", ""],
            ["", "", ""],
            ["R E C I B O","", f"VALOR: R$ {valor_sub_total_tres:.2f}"],
            ["", "", ""],
            [f"RECEBI DE VR3 LTDA. A QUANTIA DE R$: {total_pagamento['sub_total_tres']:.2f}","",""],

            ["", "", ""]
            
            
        ]
        header_table = Table(header)
        header_table.setStyle(TableStyle([('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                                           ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                                           ('ALIGN', (5, 0), (5, 0), 'LEFT'), 
                                           ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                           ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                                           ('FONTSIZE', (0, 0), (-1, -1), 14)]))
        
        elements.append(header_table)

    def _prepare_content(self, total_pagamento):
        data_inicio_formatada = (self.funcao.data_inicio).strftime('%d/%m/%Y') 
        data_fim_formatada = (self.funcao.data_fim).strftime('%d/%m/%Y')

        return [
            [f"PROVENTOS DE PRESTAÇÃO DE SERVIÇOS NO PERÍODO DE", f"{data_inicio_formatada}","Até", f"{data_fim_formatada}"],
            [f"NOME: {self.funcao.funcionario['nome_funcionario']}", "    QUAT.  VL R$","","PROVENTO"],
            [f"HORAS TRABALHADAS:", f"{self.funcao.horas_trabalhadas:.2f}  X  {self.funcao.funcionario['valor_hora_base']:.2f}","=", f"{total_pagamento['pagamento_base']:.2f}"],
            [f"REPOUSO REMUNERADO:", f"{self.funcao.repouso_remunerado:.2f}  X  {self.funcao.funcionario['repouso_remunerado']:.2f}","=", f"{total_pagamento['pagamento_folga_remunerada']:.2f}"],
            [f"HORAS EXTRAS DE 50%:", f"{self.funcao.horas_extras_um:.2f}  X  {self.funcao.funcionario['valor_hora_extra_um']:.2f}","=", f"{total_pagamento['pagamento_horas_extras_um']:.2f}"],
            [f"HORAS EXTRAS DE 100%:", f"{self.funcao.horas_extras_dois:.2f}  X  {self.funcao.funcionario['valor_hora_extra_dois']:.2f}","=", f"{total_pagamento['pagamento_horas_extras_dois']:.2f}"],
            [f"ADICIONAL NOTURNO:", f"{self.funcao.horas_noturnas:.2f}  X  {self.funcao.funcionario['adicional_noturno']:.2f}","=", f"{total_pagamento['pagamento_adicional_noturno']:.2f}"],
            [f"SUB-TOTAL 1:","","", f"{total_pagamento['sub_total_um']:.2f}",],
            [f"PAG. FÉRIAS ({self.funcao.funcionario['valor_ferias']:.2f}%):","","", f"{total_pagamento['sub_total_um_um']:.2f}", ],
            [f"PAG. 1/3 FÉRIAS ({self.funcao.funcionario['valor_um_terco_ferias']:.2f}%):", "","",f"{total_pagamento['sub_total_um_dois']:.2f}", ],
            [f"PAG. 13° SALÁRIO ({self.funcao.funcionario['valor_decimo_terceiro']:.2f}%):","","", f"{total_pagamento['sub_total_um_tres']:.2f}", ],
            [f"PAG FGTS ({self.funcao.funcionario['pagamento_fgts']:.2f}%):","","", f"{total_pagamento['sub_total_um_cinco']:.2f}", ],
            [f"SUB-TOTAL 2","","", f"{total_pagamento['sub_total_dois']:.2f}",],
            [f"PAG. INSS ({self.funcao.funcionario['desconto_inss']:.2f}%):","","", f"{total_pagamento['sub_total_dois_seis']:.2f}"],
            [f"DESC. REFEIÇÃO ({self.funcao.funcionario['desconto_refeicao']:.2f}% de Hs Trab + Repouso):","","", f"{total_pagamento['sub_total_dois_sete']:.2f}"],
            [f"DESC.TRANSPORTE ({self.funcao.funcionario['desconto_transporte']:.2f}% de Hs Trab + Repouso):","","", f"{total_pagamento['sub_total_dois_oito']:.2f}"],
            [f"COREÇÃO (+) :","","", f"{total_pagamento['sub_total_dois_nove']:.2f}"],
            [f"CORREÇÃO (-):","","", f"{total_pagamento['sub_total_dois_dez']:.2f}"],
            [f"SALDO A RECEBER:","","", f"{total_pagamento['sub_total_tres']:.2f}"],
            ["Obs: Comunicamos que providencie sua documentação completa para a realização do exame adimissional (ASO) e", "",""],
            ["Registro Legal  na Empresa.", "", ""]
            
        ]

    def _create_table(self, content):
        table = Table(content)
        table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'), # Alinhamento à esquerda
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), # Alinhamento à direita para a segunda coluna
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), # Fonte
                    ('FONTSIZE', (0, 0), (-1, -1), 12), # Tamanho da fonte
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding na parte inferior do cabeçalho   
                    ('FONTSIZE', (0, -1), (-1, -1), 12),  # Tamanho da fonte para a última linha
                    ('FONTSIZE', (0, 19), (-1, 19), 10),  # Tamanho da fonte para a linha 20 (índice 19)
                    ('FONTSIZE', (0, 20), (-1, 20), 10),  # Tamanho da fonte para a linha 21 (índice 20)
                     ('ALIGN', (0, 19), (-1, 19), 'LEFT'),  # Alinhamento à esquerda para a linha 20
                     ('SPAN', (0, 19), (3, 19)),  # Faz com que a linha 20 ocupe as quatro colunas
                    # Negrito para "SUB-TOTAL 1" e "SUB-TOTAL 2" e a primeira Linha
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrito para 1 linhas'
                    ('FONTNAME', (0, 12), (-1, 12), 'Helvetica-Bold'), # Negrito para "SUB-TOTAL 2"
                    ('FONTNAME', (0, 7), (-1, 7), 'Helvetica-Bold'),  # Negrito para "SUB-TOTAL 1"
                    ('FONTNAME', (0, 18), (-1, 18), 'Helvetica-Bold')
                
                    
                    
                ]))
            
        return table

    def _draw_footer(self, elements):

        data_pagamento = self.funcao.data_pagamento.strftime("%d/%m/%Y")
        
        footer = [
            [f"ANANINDEUA. {data_pagamento}","",""],
            ["_______________________________________________________", "", ""],
            [self.funcao.funcionario['nome_funcionario'] + f"- {self.funcao.nome_cargo}", ""],
            [f"CPF: {self.funcao.funcionario['numero_cpf']}","",""],
            [f"CHAVE PIX: {self.funcao.funcionario['chave_pix']}","",""],
        ]
        footer_table = Table(footer)
        footer_table.setStyle(TableStyle([('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                                          ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                          ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                                          ('FONTSIZE', (0, 0), (-1, -1), 14)]))
        elements.append(footer_table)
